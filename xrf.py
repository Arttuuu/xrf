import csv
import os
import glob
import shutil
import datetime
import ctypes
import itertools

import openpyxl as xl
from openpyxl.styles import Font, Alignment

# group up all the csv files in this directory to a list
csv_files = glob.glob(os.path.join(os.getcwd(), "*.csv"))
excel_file = "Tulostiedosto Sulate ver.xlsx"
excels = []
date_format = "%Y-%b-%d %X" # how CSV-file represents a date

wb = xl.load_workbook(excel_file)
ws = wb.active # define worksheet to work on
substance_row = 11 # the row where all the compounds are listed

compound_order = {} # dictionary to hold compound as key and its column number as value
for row in ws.iter_rows(min_row=substance_row, max_row=substance_row, min_col=2):
    for column, cell in enumerate(row):
        compound_order[cell.value] = column + 1

compound_order.pop(None, None) # remove trailing none-key from dict if it exists

for num, file in enumerate(csv_files):
    with open(file) as csv_file:
        """ 15 lines of code ahead that could possibly be a lot easier to implement with pandas.
            Sorts each row based on date and then rows are looped through in that order.
        """
        file_reader = csv.reader(csv_file, delimiter=',')
        dates_rows = {}
        for row_num, row in enumerate(file_reader):
            dates_rows[datetime.datetime.strptime(row[3], date_format)] = row_num+1
        csv_file.seek(0)    
        dates = []
        sorted_dates_rows = {}
        for key, value in dates_rows.items():
            dates.append(key)
        dates.sort()
        
        for date in dates:
            sorted_dates_rows[date] = dates_rows[date]    
        row_order = [item for item in sorted_dates_rows.values()]
        
        methods = [] # keep track of the methods of each row in csv
        # loop through samples (sorted by date) and csv iterable
        for (row_num, row) in zip(row_order, file_reader):
             
            extras = 1 # count of extra compounds after "sum before norm."-cell
            for col, value in enumerate(row): # loop through each value in csv row
                if col > 3 and col % 2 != 0:
                    try:
                        this_row = substance_row + row_num + 2 # +2 for the extra 2 rows under compounds
                        # set column based on compound's order in excel template
                        this_column = compound_order[row[col-1]] + 1 
                        ws.cell(row=this_row, 
                                column=this_column).value = float(value)

                    # catch all compounds not defined in the dictionary and place
                    # their values after "Sum Before Norm." cell
                    except KeyError:
                        extras += 1
                        this_column = len(compound_order) + extras
                        compound_cell = ws.cell(row = substance_row, column = this_column)
                        compound_cell.value = row[col-1]
                        compound_cell.font = Font(bold=True)
                        compound_cell.alignment = Alignment(horizontal='right')
                        ws.cell(row = this_row, 
                                column = this_column).value = float(value)
                elif col == 0:
                    ws.cell(row=5, column=2).value = value 
                    methods.append(value)

                elif col == 1:
                    ws.cell(row=substance_row + row_num + 2, 
                            column=1).value = value
                    ws.cell(row=substance_row + row_num + 2, 
                            column=1).alignment = Alignment(horizontal='right')
    
    # check for None-values and insert value to them                
    for row in ws.iter_rows(min_row=substance_row+1+2, min_col=1):
        if row[0].value:    # check that row has a sample name         
            for cell in row:
                if not cell.value:
                    cell.value = "< 0.001"
                cell.alignment = Alignment(horizontal='right')
    
    excel_name = f"test_excel_{num+1}.xlsx"             
    wb.save(excel_name)  
    wb.close()     
    excel_path = os.path.join(os.getcwd(), excel_name)
    excels.append(excel_name)             
                        
def move_file(names: list, dst: str) -> None:
    """ Moves a list of files to the destination folder
        args:
            - name: name of the file with file extension
            - dst: destination folder
    """
    for name in names:
        dir = os.path.join(os.getcwd(), dst)
        if not os.path.exists(dir):
            os.mkdir(dir)
        shutil.move(name, dir)
        
if len(methods) > 1:        
    ctypes.windll.user32.MessageBoxW(0, "More than 1 method present in CSV-file", "Warning", 0)    
#move_file(excels, 'Raportit')
#move_file(csv_files, 'CSV')
